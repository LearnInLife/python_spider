# _*_ coding:utf-8 _*_

import sys
import time
from urllib import request
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
import ssl

hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'},
       {
           'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'}]


def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_times = 0

    while True:
        url = 'https://book.douban.com/tag/' + request.quote(book_tag) + '?start=' + str((page_num * 20))
        time.sleep(np.random.rand() * 5)

        try:
            headers = hds[page_num % len(hds)]
            req = request.Request(url, headers=headers)
            with request.urlopen(req, context=context) as f:
                source_code = f.read()
            # print('Data:', source_code.decode('utf-8'))
        except Exception as e:
            print(e)
            continue

        plain_text = source_code.decode('utf-8')
        # print(plain_text)
        soup = BeautifulSoup(plain_text)
        list_soup = soup.find('ul', {'class': 'subject-list'})
        try_times += 1
        if list_soup is None and try_times < 200:
            continue
        elif list_soup is None or try_times >= 200:
            break

        for book_info in list_soup.findAll('li'):
            info = book_info.find('div', {'class': 'info'})
            title = info.find('a').text.strip().replace('\n', '').replace(' ', '')
            rating = info.find('span', {'class': 'rating_nums'}).string
            if rating:
                rating = rating.strip()
            else:
                rating = '0.0'
            people_num = info.find('span', {'class': 'pl'}).string
            if people_num:
                people_num = people_num.strip()[1:-4]
            else:
                people_num = '0'
            desc = info.find('div', {'class': 'pub'}).string.strip()
            desc_list = desc.split('/')
            try:
                author_info = '作者/译者: ' + '/'.join(desc_list[0:-3])
            except Exception:
                author_info = '作者/译者: 暂无'

            try:
                pub_info = '出版信息: ' + '/'.join(desc_list[-3:-1])
            except Exception:
                pub_info = '出版信息: 暂无'

            try:
                price = desc_list[-1]
            except Exception:
                price = '暂无'
            # title = str(title)
            # people_num = people_num.decode('utf-8')
            book = [title, rating, people_num, author_info, pub_info,price]
            book_list.append(book)
            try_times = 0
        page_num += 1
        if page_num >= 5:
            break
    return book_list


def print_book_excel(book_list, tag_lists):
    wb = Workbook()
    ws = []
    for i in range(len(tag_lists)):
        ws.append(wb.create_sheet(title=tag_lists[i]))
    for i in range(len(tag_lists)):
        ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社','价格'])
        count = 1
        for bl in book_list[i]:
            a = [count, bl[0], bl[1], bl[2], bl[3], bl[4],bl[5]]
            print(a)
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4],bl[5]])
            count += 1
    save_path = 'book_list'
    for i in range(len(tag_lists)):
        save_path += ('-' + tag_lists[i])
    save_path += '.xlsx'
    wb.save(save_path)


def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists


context = ssl._create_unverified_context()

if __name__ == '__main__':
    book_tag_list = ['小说', '编程', '散文']
    book = do_spider(book_tag_list)
    print_book_excel(book, book_tag_list)
